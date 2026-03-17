"""
collector/excel_reader.py
-------------------------
Reads production-planning data from Excel files and caches the results in
SQLite to avoid re-parsing on every dashboard refresh.

All functions follow the same pattern:

1. Compare the file's mtime against the ``excel_cache`` table in *sqlite_conn*.
2. Return cached data immediately when the mtime matches.
3. Otherwise, copy the workbook to a temp file (strips VBA macros), parse it
   with openpyxl, then persist the result back to the cache.

Provides
--------
- ``load_ploegen_definitions(excel_path, sqlite_conn)``
- ``load_vsm_data(excel_path, sqlite_conn)``
- ``load_planning_data(excel_path, sqlite_conn)``
- ``load_otif_data(excel_path, sqlite_conn)``
"""

import glob
import json
import os
import tempfile
import traceback
from collections import defaultdict
from datetime import datetime, date


# ---------------------------------------------------------------------------
# Internal SQLite cache helpers
# ---------------------------------------------------------------------------

def _excel_cache_create_table(sqlite_conn) -> None:
    """Ensure the ``excel_cache`` table exists in *sqlite_conn*."""
    try:
        sqlite_conn.execute('''
            CREATE TABLE IF NOT EXISTS excel_cache (
                cache_key TEXT PRIMARY KEY,
                mtime     REAL NOT NULL,
                data_json TEXT NOT NULL,
                saved_at  TEXT NOT NULL
            )
        ''')
        sqlite_conn.commit()
    except Exception:
        pass


def _excel_cache_load(sqlite_conn, cache_key: str, current_mtime: float):
    """Return deserialised data dict if SQLite has a matching mtime entry, else ``None``.

    Parameters
    ----------
    sqlite_conn : sqlite3.Connection
        An open SQLite connection that contains the ``excel_cache`` table.
    cache_key : str
        Logical name for the cached dataset (e.g. ``'ploegen'``, ``'vsm'``).
    current_mtime : float
        The file's current modification time (from ``os.path.getmtime``).
    """
    if sqlite_conn is None:
        return None
    try:
        row = sqlite_conn.execute(
            'SELECT mtime, data_json FROM excel_cache WHERE cache_key = ?',
            (cache_key,)
        ).fetchone()
        if row and abs(row[0] - current_mtime) < 0.001:
            return json.loads(row[1])
    except (json.JSONDecodeError, Exception) as e:
        print(f"[WARNING] Excel cache load failed for {cache_key!r}: {e}")
    return None


def _excel_cache_save(sqlite_conn, cache_key: str, mtime: float, data) -> None:
    """Serialise *data* as JSON and upsert into the ``excel_cache`` table.

    Parameters
    ----------
    sqlite_conn : sqlite3.Connection
        An open SQLite connection.
    cache_key : str
        Logical name for the cached dataset.
    mtime : float
        The file's modification time at the time of parsing.
    data : any
        JSON-serialisable object to cache.
    """
    if sqlite_conn is None:
        return
    try:
        sqlite_conn.execute(
            'INSERT OR REPLACE INTO excel_cache (cache_key, mtime, data_json, saved_at) '
            'VALUES (?, ?, ?, ?)',
            (cache_key, mtime, json.dumps(data), datetime.now().isoformat())
        )
        sqlite_conn.commit()
    except Exception as e:
        print(f"[WARNING] Excel cache save failed for {cache_key!r}: {e}")


# ---------------------------------------------------------------------------
# Public loaders
# ---------------------------------------------------------------------------

def load_ploegen_definitions(excel_path: str, sqlite_conn) -> tuple:
    """Load ploeg definitions from *Ploegen.xlsx*.

    Reads ploeg member names from columns L–P (rows 1–99) and the
    *ploegenwissel* date from cell Q1.  Results are cached in SQLite keyed
    by the file's mtime so subsequent calls return immediately when the file
    has not changed.

    Parameters
    ----------
    excel_path : str
        Filesystem path to ``Ploegen.xlsx``.
    sqlite_conn : sqlite3.Connection or None
        An open SQLite connection used for mtime-based caching.  Pass
        ``None`` to disable caching (the file is always re-parsed).

    Returns
    -------
    tuple[dict, date | None]
        ``(ploegen_data, ploegenwissel_date)`` where *ploegen_data* maps
        first-two-letter abbreviations to ploeg dicts, and
        *ploegenwissel_date* is the date of the last shift rotation (or
        ``None`` if the cell is absent/unparseable).
    """
    import openpyxl

    temp_file = None

    # Ensure cache table exists
    if sqlite_conn:
        _excel_cache_create_table(sqlite_conn)

    # mtime-based cache check
    try:
        current_mtime = os.path.getmtime(excel_path)
        sq = _excel_cache_load(sqlite_conn, 'ploegen', current_mtime)
        if sq:
            ploegen_data = sq['ploegen_data']
            pwd = sq.get('ploegenwissel_date')
            ploegenwissel_date = date.fromisoformat(pwd) if pwd else None
            print("  (ploegen excel unchanged — using SQLite cache)")
            return ploegen_data, ploegenwissel_date
    except Exception:
        pass  # If mtime check fails, proceed with normal load

    try:
        # Create temporary macro-free copy
        wb_original = openpyxl.load_workbook(excel_path, data_only=True)

        temp_fd, temp_file = tempfile.mkstemp(suffix='.xlsx')
        os.close(temp_fd)
        wb_original.save(temp_file)
        wb_original.close()

        # Read from the clean temp file
        wb = openpyxl.load_workbook(temp_file, data_only=True)
        sheet = wb.active

        # Read ploegenwissel date from cell Q1
        ploegenwissel_cell = sheet['Q1'].value
        if ploegenwissel_cell:
            if isinstance(ploegenwissel_cell, datetime):
                ploegenwissel_date = ploegenwissel_cell.date()
            elif isinstance(ploegenwissel_cell, date):
                ploegenwissel_date = ploegenwissel_cell
            else:
                try:
                    ploegenwissel_date = datetime.strptime(
                        str(ploegenwissel_cell), '%Y-%m-%d'
                    ).date()
                except Exception:
                    ploegenwissel_date = None
                    print(f"Warning: Could not parse ploegenwissel date from Q1: {ploegenwissel_cell}")
        else:
            ploegenwissel_date = None

        if not ploegenwissel_date:
            print("Warning: No ploegenwissel date found in Q1, using all historical data")

        ploegen = {}
        for col_idx, col_letter in enumerate(['L', 'M', 'N', 'O', 'P'], start=1):
            names = []
            for row in range(1, 100):  # START FROM ROW 1 (not 2!)
                cell_value = sheet[f'{col_letter}{row}'].value
                if cell_value and str(cell_value).strip():
                    names.append(str(cell_value).strip())
                elif names:  # Stop at first empty after finding names
                    break

            if names:
                ploeg_name = ", ".join(names)
                for name in names:
                    first_two = name[:2].upper()
                    ploegen[first_two] = {
                        'ploeg_number': col_idx,
                        'ploeg_name':   ploeg_name,
                        'members':      names,
                    }

        wb.close()

        # Persist to SQLite cache
        try:
            mtime_now = os.path.getmtime(excel_path)
            _excel_cache_save(sqlite_conn, 'ploegen', mtime_now, {
                'ploegen_data':      ploegen,
                'ploegenwissel_date': ploegenwissel_date.isoformat() if ploegenwissel_date else None,
            })
        except Exception:
            pass

        return ploegen, ploegenwissel_date

    except Exception as e:
        print(f"Warning: Could not load ploegen data: {e}")
        return {}, None
    finally:
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except Exception:
                pass


def load_vsm_data(excel_path: str, sqlite_conn) -> dict:
    """Load VSM (daily management) Excel data with mtime-based SQLite caching.

    Reads the active sheet, interprets columns B–K as isotope targetstroom
    values, and computes an OTIF% per row and a week average.

    Parameters
    ----------
    excel_path : str
        Filesystem path to the VSM ``.xlsx`` file.
    sqlite_conn : sqlite3.Connection or None
        An open SQLite connection used for mtime-based caching.

    Returns
    -------
    dict or list
        On success: ``{'rows': [...], 'week_avg_otif': float|None}``.
        On failure or if *excel_path* is falsy: an empty list ``[]``.
    """
    import openpyxl

    if not excel_path:
        return []

    # Ensure cache table exists
    if sqlite_conn:
        _excel_cache_create_table(sqlite_conn)

    # mtime-based cache check
    try:
        current_mtime = os.path.getmtime(excel_path)
        sq = _excel_cache_load(sqlite_conn, 'vsm', current_mtime)
        if sq:
            print("  (VSM excel unchanged — using SQLite cache)")
            return sq['vsm_data']
    except Exception:
        pass

    temp_file = None

    # Targets per isotope group used for OTIF% computation
    TARGETS = {
        'tl':   170.0,   # Tl1-Tl5 (cols B-F, indices 1-5)
        'ga':    80.0,   # Ga      (col G, index 6)
        'in':    80.0,   # In      (col H, index 7)
        'rb':    70.0,   # Rb x2   (cols I-J, indices 8-9)
        'i123': 100.0,   # I123    (col K, index 10)
    }

    try:
        wb_original = openpyxl.load_workbook(excel_path, data_only=True)
        temp_fd, temp_file = tempfile.mkstemp(suffix='.xlsx')
        os.close(temp_fd)
        wb_original.save(temp_file)
        wb_original.close()

        wb = openpyxl.load_workbook(temp_file, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # Row 0 = header (B1:K1 = Tl1..I123, M1 = OTIF)
        headers = list(rows[0])  # 14 cols (A..N)

        data_rows = []
        for row in rows[1:]:
            date_val = row[0] if len(row) > 0 else None
            if date_val is None:
                continue

            # Format date
            if hasattr(date_val, 'strftime'):
                date_str = date_val.strftime('%d-%m-%Y')
            else:
                date_str = str(date_val)

            # Numeric values per column (B=1 .. K=10, M=12)
            def num(idx):
                if len(row) > idx and row[idx] is not None:
                    try:
                        return float(row[idx])
                    except Exception:
                        pass
                return None

            tl_vals = [num(i) for i in range(1, 6)]   # Tl1-Tl5
            ga_val  = num(6)
            in_val  = num(7)
            rb_vals = [num(8), num(9)]                  # Rb x2
            i123    = num(10)

            # Compute OTIF% — same logic as the Excel formula
            def group_otif(vals, target):
                valid = [v for v in vals if v is not None]
                if not valid:
                    return None, False
                return sum(valid) / len(valid) / target * 100, True

            tl_pct,   tl_ok   = group_otif(tl_vals,   TARGETS['tl'])
            ga_pct,   ga_ok   = group_otif([ga_val],   TARGETS['ga'])
            in_pct,   in_ok   = group_otif([in_val],   TARGETS['in'])
            rb_pct,   rb_ok   = group_otif(rb_vals,    TARGETS['rb'])
            i123_pct, i123_ok = group_otif([i123],     TARGETS['i123'])

            active = sum([tl_ok, ga_ok, in_ok, rb_ok, i123_ok])
            if active > 0:
                otif_pct = round((
                    (tl_pct   if tl_ok   else 0) +
                    (ga_pct   if ga_ok   else 0) +
                    (in_pct   if in_ok   else 0) +
                    (rb_pct   if rb_ok   else 0) +
                    (i123_pct if i123_ok else 0)
                ) / active, 1)
            else:
                otif_pct = None

            data_rows.append({
                'date': date_str,
                'tl1':  num(1),
                'tl2':  num(2),
                'tl3':  num(3),
                'tl4':  num(4),
                'tl5':  num(5),
                'ga':   ga_val,
                'in_':  in_val,
                'rb1':  num(8),
                'rb2':  num(9),
                'i123': i123,
                'otif': otif_pct,
            })

        # Week-average OTIF across all parsed rows
        otif_values = [r['otif'] for r in data_rows if r['otif'] is not None]
        week_avg_otif = round(sum(otif_values) / len(otif_values), 1) if otif_values else None

        vsm_data = {
            'rows':          data_rows,
            'week_avg_otif': week_avg_otif,
        }

        wb.close()

        # Persist to SQLite cache
        try:
            mtime_now = os.path.getmtime(excel_path)
            _excel_cache_save(sqlite_conn, 'vsm', mtime_now, {'vsm_data': vsm_data})
        except Exception:
            pass

        print(f"Loaded VSM data: {len(data_rows)} rows")
        return vsm_data

    except Exception as e:
        print(f"Warning: Could not load VSM data: {e}")
        return []
    finally:
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except Exception:
                pass


def load_planning_data(excel_path: str, sqlite_conn) -> dict:
    """Load planning data from *Planning & Control Cyclotron.xlsm*.

    Reads the *Verlof* sheet starting at row 7.  For each date row the
    function picks the first non-empty lead-person name from each shift's
    column range:

    * Ochtenddienst: F, G, H, I
    * Middagdienst:  J, K, L, M
    * Nachtdienst:   Q, R, S, T

    Results are cached in SQLite keyed by the file's mtime.

    Parameters
    ----------
    excel_path : str
        Filesystem path to the ``.xlsm`` planning file.
    sqlite_conn : sqlite3.Connection or None
        An open SQLite connection used for mtime-based caching.

    Returns
    -------
    dict[date, dict]
        Maps each date to ``{'ochtenddienst': str|None, 'middagdienst': str|None,
        'nachtdienst': str|None}``.  Returns ``{}`` on failure.
    """
    import openpyxl

    temp_file = None

    # Ensure cache table exists
    if sqlite_conn:
        _excel_cache_create_table(sqlite_conn)

    # mtime-based cache check
    try:
        current_mtime = os.path.getmtime(excel_path)
        sq = _excel_cache_load(sqlite_conn, 'planning', current_mtime)
        if sq:
            planning_data = {
                date.fromisoformat(k): v for k, v in sq['planning_data'].items()
            }
            print("  (planning excel unchanged — using SQLite cache)")
            return planning_data
    except Exception:
        pass  # If mtime check fails, proceed with normal load

    try:
        # Create temporary macro-free copy
        wb_original = openpyxl.load_workbook(excel_path, data_only=True)

        temp_fd, temp_file = tempfile.mkstemp(suffix='.xlsx')
        os.close(temp_fd)
        wb_original.save(temp_file)
        wb_original.close()

        # Read from the clean temp file
        wb = openpyxl.load_workbook(temp_file, data_only=True)
        sheet = wb['Verlof']

        planning = {}
        for row in range(7, 5000):  # Start at row 7, read up to 5000
            date_val = sheet[f'B{row}'].value

            if date_val is None:
                break

            # Convert to date
            if isinstance(date_val, datetime):
                row_date = date_val.date()
            elif isinstance(date_val, str):
                try:
                    row_date = datetime.strptime(date_val, '%Y-%m-%d').date()
                except Exception:
                    continue
            else:
                continue

            # Extract lead person for each shift — check multiple columns
            # Ochtenddienst: F, G, H, I
            ochtend_lead = None
            for col in ['F', 'G', 'H', 'I']:
                val = sheet[f'{col}{row}'].value
                if val and str(val).strip():
                    ochtend_lead = str(val).strip()
                    break

            # Middagdienst: J, K, L, M
            middag_lead = None
            for col in ['J', 'K', 'L', 'M']:
                val = sheet[f'{col}{row}'].value
                if val and str(val).strip():
                    middag_lead = str(val).strip()
                    break

            # Nachtdienst: Q, R, S, T
            nacht_lead = None
            for col in ['Q', 'R', 'S', 'T']:
                val = sheet[f'{col}{row}'].value
                if val and str(val).strip():
                    nacht_lead = str(val).strip()
                    break

            planning[row_date] = {
                'ochtenddienst': ochtend_lead,
                'middagdienst':  middag_lead,
                'nachtdienst':   nacht_lead,
            }

        wb.close()

        # Persist to SQLite cache — date keys serialised as ISO strings for JSON
        try:
            mtime_now = os.path.getmtime(excel_path)
            _excel_cache_save(sqlite_conn, 'planning', mtime_now, {
                'planning_data': {k.isoformat(): v for k, v in planning.items()},
            })
        except Exception:
            pass

        return planning

    except Exception as e:
        print(f"Warning: Could not load planning data: {e}")
        traceback.print_exc()
        return {}
    finally:
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except Exception:
                pass


def load_otif_data(sqlite_conn) -> tuple:
    """Load OTIF KPI and missed-orders data from the weekly OTIF Excel file.

    Searches ``X:\\Cyclotron Bureau\\OTIF`` for a file matching
    ``OTIF week*.xlsx`` and parses two sheets:

    * **7 KPI Week** — weekly OTIF percentages per isotope (last 52 weeks).
    * **1Log** — individual missed-order records filtered to known products
      and reasons.

    Results are cached in SQLite keyed by the file's mtime.

    Parameters
    ----------
    sqlite_conn : sqlite3.Connection or None
        An open SQLite connection used for mtime-based caching.

    Returns
    -------
    tuple[list, dict]
        ``(otif_kpi_data, otif_table_data)`` where *otif_kpi_data* is a list
        of weekly KPI dicts and *otif_table_data* is a nested dict
        ``{product: {week_label: missed_count}}``.
        Returns ``([], {})`` on failure or if no OTIF file is found.
    """
    import openpyxl

    otif_folder = r"X:\Cyclotron Bureau\OTIF"
    pattern = os.path.join(otif_folder, "OTIF week*.xlsx")
    matches = glob.glob(pattern)
    if not matches:
        print(f"Warning: No OTIF file found in {otif_folder}")
        return [], {}

    otif_path = matches[0]
    temp_file = None

    # Ensure cache table exists
    if sqlite_conn:
        _excel_cache_create_table(sqlite_conn)

    # mtime-based cache check
    try:
        current_mtime = os.path.getmtime(otif_path)
        sq = _excel_cache_load(sqlite_conn, 'otif', current_mtime)
        if sq:
            print("  (OTIF excel unchanged — using SQLite cache)")
            return sq['otif_kpi_data'], sq['otif_table_data']
    except Exception:
        pass  # If mtime check fails, proceed with normal load

    try:
        wb_original = openpyxl.load_workbook(otif_path, data_only=True)
        temp_fd, temp_file = tempfile.mkstemp(suffix='.xlsx')
        os.close(temp_fd)
        wb_original.save(temp_file)
        wb_original.close()

        wb = openpyxl.load_workbook(temp_file, data_only=True)

        # ── 7 KPI Week: columns A, G, H, J, K, L ─────────────────────────────
        kpi_sheet = wb['7 KPI Week']
        today_iso  = datetime.now().isocalendar()
        today_year, today_week = today_iso[0], today_iso[1]

        def week_label_to_tuple(label):
            try:
                y, w = str(label).split('W')
                return (int(y), int(w))
            except Exception:
                return None

        def normalize_pct(v):
            if v is None:
                return None
            try:
                f = float(v)
                # Stored as 0-1 decimal → convert to 0-100
                if 0.0 <= f <= 1.0:
                    return round(f * 100, 1)
                return round(f, 1)
            except Exception:
                return None

        kpi_data = []
        for row in kpi_sheet.iter_rows(min_row=2, values_only=True):
            week_label = row[0]   # Col A
            if not week_label:
                continue
            tup = week_label_to_tuple(week_label)
            if not tup:
                continue
            year, week = tup
            # Keep only last 52 weeks
            weeks_diff = (today_year * 53 + today_week) - (year * 53 + week)
            if weeks_diff < 0 or weeks_diff > 52:
                continue
            kpi_data.append({
                'week':             str(week_label),
                'year':             year,
                'isoweek':          week,
                'gallium':          normalize_pct(row[6]  if len(row) > 6  else None),  # Col G
                'i123':             normalize_pct(row[7]  if len(row) > 7  else None),  # Col H
                'indium':           normalize_pct(row[9]  if len(row) > 9  else None),  # Col J
                'thallium':         normalize_pct(row[10] if len(row) > 10 else None),  # Col K
                'rubidium_krypton': normalize_pct(row[11] if len(row) > 11 else None),  # Col L
            })

        kpi_data.sort(key=lambda x: (x['year'], x['isoweek']))

        # ── 1Log: header row 2, data from row 3 ──────────────────────────────
        # Col D (idx 3) = week, G (idx 6) = OTIF count, I (idx 8) = product,
        # K (idx 10) = reason
        log_sheet = wb['1Log']
        PRODUCT_FILTER = {'Gallium', 'I-123', 'Indium', 'Krypton', 'Thallium'}
        REASON_FILTER  = {'Production', 'Equipment failure Petten',
                          'Bulk shortage', 'Production Petten'}

        otif_table = defaultdict(lambda: defaultdict(int))
        for row in log_sheet.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 11:
                continue
            week_val = row[3]    # Col D
            otif_val = row[6]    # Col G
            product  = row[8]    # Col I
            reason   = row[10]   # Col K

            if not product or str(product).strip() not in PRODUCT_FILTER:
                continue
            if not reason or str(reason).strip() not in REASON_FILTER:
                continue
            if not week_val:
                continue

            try:
                val = int(otif_val) if otif_val is not None else 0
            except Exception:
                val = 0

            otif_table[str(product).strip()][str(week_val).strip()] += val

        otif_table_data = {k: dict(v) for k, v in otif_table.items()}
        wb.close()

        # Persist to SQLite cache
        try:
            mtime_now = os.path.getmtime(otif_path)
            _excel_cache_save(sqlite_conn, 'otif', mtime_now, {
                'otif_kpi_data':   kpi_data,
                'otif_table_data': otif_table_data,
            })
        except Exception:
            pass

        n_entries = sum(len(v) for v in otif_table_data.values())
        print(f"Loaded OTIF data: {len(kpi_data)} KPI weeks, {n_entries} missed-order entries")
        return kpi_data, otif_table_data

    except Exception as e:
        print(f"Warning: Could not load OTIF data: {e}")
        traceback.print_exc()
        return [], {}
    finally:
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except Exception:
                pass
